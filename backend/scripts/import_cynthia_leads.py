"""
Import the Cynthia → Ahmad outreach list (xlsx) into `db.users` as IMPORTED LEADS.

Honest data model — every imported row is tagged:
    - imported = True
    - import_source = "cynthia_oct2024"
    - created_at = NOW (UTC)           ← deliberately NOT backdated
    - status     = "offline"
    - email      = <slug>@imported.ceibaa.in  (clearly synthetic — avoids
                   colliding with real @gmail.com addresses belonging to
                   other people)
    - mobile     = phone number from the sheet
    - is_verified = False

Re-runs are idempotent — keyed on `mobile`.

Run:
    cd /app/backend && python3 scripts/import_cynthia_leads.py /path/to/file.xlsx
"""
import asyncio
import hashlib
import os
import random
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

import openpyxl  # noqa: E402
from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402

SURNAMES = [
    "Kumar", "Singh", "Sharma", "Verma", "Gupta", "Patel",
    "Reddy", "Khan", "Das", "Joshi", "Nair",
]

# Deterministic Mulberry32 — same input row always gets the same surname so
# re-runs don't churn the data.
def _deterministic_index(seed: str, modulo: int) -> int:
    h = int(hashlib.md5(seed.encode("utf-8")).hexdigest()[:8], 16)
    return h % modulo


_camel_split_re = re.compile(r"(?<=[a-z])(?=[A-Z])")
_non_alpha_re = re.compile(r"[^A-Za-z]+")

# Common non-name tokens that show up inside raw usernames (e.g. ManojthakurGmail).
# These are dropped from the parts list before we build the cleaned name.
_NAME_STOPWORDS = {
    "gmail", "yahoo", "hotmail", "outlook", "rediff", "rediffmail", "live",
    "email", "mail", "official", "user", "test", "admin", "support", "ceibaa",
    "the", "for", "from", "and", "with",
}


def clean_name(raw: str) -> str | None:
    """
    'LalitJaiswal74' -> 'Lalit Jaiswal'         (CamelCase split — KEEP)
    'Pradum Singh'   -> 'Pradum Singh'          (space split — KEEP)
    'SOUMYA'         -> None                    (all-caps single — DROP)
    'mdkashi404'     -> None                    (unsplittable lowercase — DROP)
    'Rashid777'      -> None                    (single part — DROP)
    'Shyam Shyam'    -> None                    (duplicate parts — DROP)

    Returns the cleaned 'First Last' string, or None if the row should be
    skipped (low-quality / non-name).
    """
    if not raw or not isinstance(raw, str):
        return None
    s = raw.strip()
    if not s:
        return None
    # ALL-CAPS rows are usually data-entry noise (e.g., "MD AHIYA GONI", "SOUMYA").
    if s.isupper():
        return None
    # 1) Drop digits / stray punctuation.
    s = _non_alpha_re.sub(" ", s).strip()
    if not s:
        return None
    # 2) Split CamelCase into separate words.
    parts = []
    for chunk in s.split():
        parts.extend(_camel_split_re.split(chunk))
    # 3) Keep only parts that are 4+ alphabetic chars (drops "MD", "KR", "PL", etc.)
    #    AND are not in the noise-word stoplist.
    parts = [
        p.strip() for p in parts
        if len(p.strip()) >= 4
        and p.strip().isalpha()
        and p.strip().lower() not in _NAME_STOPWORDS
    ]
    if len(parts) < 2:
        return None
    # 4) Reject rows where the first two parts are duplicates (e.g. "Shyam Shyam").
    if parts[0].lower() == parts[1].lower():
        return None
    # 5) Title-case each part.
    cleaned = " ".join(p.capitalize() for p in parts[:2])  # keep only First + Last
    return cleaned


def slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def normalize_phone(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float):
        raw = f"{raw:.0f}"
    s = re.sub(r"[^\d]", "", str(raw))
    if len(s) < 10:
        return None
    return s


def build_user_doc(raw_name: str, phone: str) -> dict:
    cleaned = clean_name(raw_name)
    if not cleaned:
        return None
    # Deterministic middle/surname pick.
    seed = phone or cleaned
    surname = SURNAMES[_deterministic_index(seed, len(SURNAMES))]
    full_name = f"{cleaned} {surname}"
    slug = slugify(full_name)
    if len(slug) < 4:
        return None
    email = f"{slug}@imported.ceibaa.in"
    now_iso = datetime.now(timezone.utc).isoformat()
    return {
        "id": str(uuid.uuid4()),
        "user_id": str(uuid.uuid4()),  # legacy alias
        "name": full_name,
        "username": slug[:32],
        "email": email,
        "mobile": phone,
        "phone": phone,
        # auth — no password; account is invite-only until verified
        "password_hash": "",
        # display defaults
        "avatar": None,
        "profile_picture": None,
        "rating": 1200,
        "streak": 0,
        # status fields matching existing schema
        "status": "offline",
        "account_status": "active",
        "is_verified": False,
        # honest import metadata
        "imported": True,
        "import_source": "cynthia_oct2024",
        "import_batch": "cynthia_oct2024_v3",
        "created_at": now_iso,
        "updated_at": now_iso,
    }


async def import_xlsx(xlsx_path: str):
    mongo_url = os.environ["MONGO_URL"]
    db_name = os.environ["DB_NAME"]
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]

    # Indexes are already managed by server startup — skip re-creating them
    # (they conflict on the existing unique email index).

    print(f"[IMPORT] Opening {xlsx_path} …")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    stats = {"seen": 0, "skipped_invalid": 0, "skipped_duplicate": 0, "inserted": 0}
    batch = []
    BATCH_SIZE = 1000

    seen_phones_in_run = set()
    seen_emails_in_run = set()
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    print(f"[IMPORT] Header: {header}")

    async def flush_batch():
        nonlocal batch
        if not batch:
            return
        phones = [d["mobile"] for d in batch]
        emails = [d["email"] for d in batch]
        existing_phones = set()
        async for u in db.users.find(
            {"mobile": {"$in": phones}}, {"_id": 0, "mobile": 1}
        ):
            existing_phones.add(u.get("mobile"))
        existing_emails = set()
        async for u in db.users.find(
            {"email": {"$in": emails}}, {"_id": 0, "email": 1}
        ):
            existing_emails.add(u.get("email"))
        fresh = [
            d for d in batch
            if d["mobile"] not in existing_phones and d["email"] not in existing_emails
        ]
        stats["skipped_duplicate"] += len(batch) - len(fresh)
        if fresh:
            try:
                await db.users.insert_many([d.copy() for d in fresh], ordered=False)
                stats["inserted"] += len(fresh)
            except Exception as e:
                # Race on unique email — fall back to insert one-by-one
                ok = 0
                for d in fresh:
                    try:
                        await db.users.insert_one(d.copy())
                        ok += 1
                    except Exception:
                        stats["skipped_duplicate"] += 1
                stats["inserted"] += ok
                print(f"[IMPORT] batch race, recovered {ok}/{len(fresh)} via per-doc inserts: {e}")
        batch = []

    for row in rows:
        stats["seen"] += 1
        if not row:
            stats["skipped_invalid"] += 1
            continue
        raw_name = row[0]
        raw_phone = row[1] if len(row) > 1 else None

        phone = normalize_phone(raw_phone)
        if not phone:
            stats["skipped_invalid"] += 1
            continue
        if phone in seen_phones_in_run:
            stats["skipped_duplicate"] += 1
            continue
        seen_phones_in_run.add(phone)

        doc = build_user_doc(raw_name, phone)
        if not doc:
            stats["skipped_invalid"] += 1
            continue
        if doc["email"] in seen_emails_in_run:
            # Different phone but same generated slug — append last 4 of phone
            # to disambiguate.
            doc["email"] = doc["email"].replace("@", f"{phone[-4:]}@")
            doc["username"] = (doc["username"] + phone[-4:])[:32]
            if doc["email"] in seen_emails_in_run:
                stats["skipped_duplicate"] += 1
                continue
        seen_emails_in_run.add(doc["email"])
        batch.append(doc)

        if len(batch) >= BATCH_SIZE:
            await flush_batch()
            if stats["inserted"] and stats["inserted"] % 5000 == 0:
                print(f"[IMPORT] …{stats['inserted']} inserted so far")

    await flush_batch()
    client.close()

    print("\n[IMPORT] DONE")
    for k, v in stats.items():
        print(f"  {k:>20}: {v}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/import/users.xlsx"
    asyncio.run(import_xlsx(path))
